"""Tests for SQLite history persistence and price-variation detection."""
from openpyxl import load_workbook

from pricewatch.models.change import compute_changes
from pricewatch.models.product import Product
from pricewatch.services.history import HistoryStore
from pricewatch.services.report import ReportBuilder

URL = "https://books.toscrape.com/catalogue/book_{}/index.html"


def _book(i: int, price: float) -> Product:
    return Product(title=f"Book {i}", price=price, rating=(i % 5) + 1,
                   in_stock=True, url=URL.format(i))


# ---- pure variation logic -----------------------------------------------
def test_compute_changes_classifies_trends():
    products = [_book(1, 12.0), _book(2, 8.0), _book(3, 5.0), _book(4, 9.0)]
    previous = {URL.format(1): 10.0, URL.format(2): 10.0, URL.format(3): 5.0}

    by_title = {c.product.title: c for c in compute_changes(products, previous)}

    up = by_title["Book 1"]
    assert up.trend == "up" and up.delta == 2.0 and up.pct == 20.0

    down = by_title["Book 2"]
    assert down.trend == "down" and down.delta == -2.0 and down.pct == -20.0

    same = by_title["Book 3"]
    assert same.trend == "same" and same.delta == 0.0

    new = by_title["Book 4"]  # no previous price recorded
    assert new.trend == "new" and new.delta is None and new.pct is None


# ---- SQLite round-trip ---------------------------------------------------
def test_history_store_records_and_diffs(tmp_path):
    db = tmp_path / "history.db"

    with HistoryStore(db) as store:
        # First run: everything is new, nothing to compare against.
        first = store.diff_and_record([_book(1, 10.0), _book(2, 20.0)])
        assert {c.trend for c in first} == {"new"}

    with HistoryStore(db) as store:
        # Second run on the *same* file: prices compared to the first run.
        second = store.diff_and_record([_book(1, 11.0), _book(2, 18.0)])
        trends = {c.product.title: c.trend for c in second}
        assert trends == {"Book 1": "up", "Book 2": "down"}

        # latest_prices now reflects the most recent run.
        assert store.latest_prices()[URL.format(1)] == 11.0


# ---- report integration --------------------------------------------------
def test_report_includes_changes_sheet(tmp_path):
    products = [_book(1, 11.0), _book(2, 18.0)]
    changes = compute_changes(products, {URL.format(1): 10.0, URL.format(2): 20.0})

    path = ReportBuilder(tmp_path).build(products, changes)

    wb = load_workbook(path)
    assert {"Data", "Summary", "Changes"} <= set(wb.sheetnames)
    # Both products moved, so they appear in the Changes sheet (header + 2 rows).
    assert wb["Changes"].max_row == 3
