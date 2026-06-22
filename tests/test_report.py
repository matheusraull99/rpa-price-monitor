"""Integration test: a report file is produced with the expected sheets."""
from openpyxl import load_workbook

from pricewatch.models.product import Product
from pricewatch.services.report import ReportBuilder

SAMPLE = [
    Product(title=f"Book {i}", price=10.0 + i, rating=(i % 5) + 1, in_stock=bool(i % 2),
            url=f"https://books.toscrape.com/catalogue/book-{i}/index.html")
    for i in range(6)
]


def test_report_is_created(tmp_path):
    path = ReportBuilder(tmp_path).build(SAMPLE)
    assert path.exists() and path.suffix == ".xlsx"
    wb = load_workbook(path)
    assert {"Data", "Summary"} <= set(wb.sheetnames)
    assert wb["Data"].max_row == len(SAMPLE) + 1  # header + rows
